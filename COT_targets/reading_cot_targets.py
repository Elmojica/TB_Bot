import openpyxl, pyautogui, time

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("COT.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

search_coordinates_x, search_coordinates_y = pyautogui.position(87, 822)
enter_x_coord_x, enter_x_coord_y = pyautogui.position(953, 545)
enter_y_coord_x, enter_y_coord_y = pyautogui.position(1065, 545)
go_to_location_x, go_to_location_y = pyautogui.position(955, 587)
select_city_x, select_city_y = pyautogui.position(957, 568)
# create_landmark_x, create_landmark_y = pyautogui.position(1018, 424)
# enter_description_x, enter_description_y = pyautogui.position(774, 499)
# save_landmark_x, save_landmark_y = pyautogui.position(954, 569)
copy_location_x, copy_location_y = pyautogui.position(1044, 423)
click_chat_x, click_chat_y = pyautogui.position(1048, 1013)
search_player_chat_x, search_player_chat_y = pyautogui.position(602, 355)
select_chat_x, select_chat_y = pyautogui.position(633, 433)
paste_location_x, paste_location_y = pyautogui.position(941, 855)
confirm_paste_x, confirm_paste_y = pyautogui.position(870, 635)
click_out_x, click_out_y = pyautogui.position(1567, 555)# double click on this one
chat_input_x, chat_input_y = pyautogui.position(990, 802)
send_message_x, send_message_y = pyautogui.position(870, 854)
time.sleep(3)
def transfer_coords(x_coord, y_coord, clannie):
    #Search target
    pyautogui.moveTo(search_coordinates_x, search_coordinates_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(enter_x_coord_x, enter_x_coord_y, 0.5)
    time.sleep(0.5)
    pyautogui.click()
    pyautogui.typewrite(x_coord, interval=0.1)
    pyautogui.moveTo(enter_y_coord_x, enter_y_coord_y, 0.5)
    time.sleep(0.5)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.typewrite(y_coord, interval=0.1)

    #getting target
    pyautogui.moveTo(go_to_location_x, go_to_location_y, 0.5)
    pyautogui.click()
    time.sleep(3)
    pyautogui.moveTo(select_city_x, select_city_y, 0.5)
    pyautogui.click()

    #sending target
    pyautogui.moveTo(copy_location_x, copy_location_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(click_out_x, click_out_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(click_chat_x, click_chat_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(search_player_chat_x, search_player_chat_y, 0.5)
    pyautogui.click()
    pyautogui.typewrite(clannie, interval=0.1)
    pyautogui.moveTo(select_chat_x, select_chat_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(paste_location_x, paste_location_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(confirm_paste_x, confirm_paste_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(click_out_x, click_out_y, 0.5)
    pyautogui.click()
    # pyautogui.click()

def send_label(label): 
    pyautogui.moveTo(click_chat_x, click_chat_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(search_player_chat_x, search_player_chat_y, 0.5)
    pyautogui.click()
    pyautogui.typewrite(clannie)
    pyautogui.moveTo(select_chat_x, select_chat_y, 0.5)
    pyautogui.click()
    pyautogui.moveTo(chat_input_x, chat_input_y, 0.5)
    pyautogui.click()
    pyautogui.write(label)
    pyautogui.moveTo(send_message_x, send_message_y, 0.5)
    # pyautogui.click()
    pyautogui.moveTo(click_out_x, click_out_y, 0.5)
    pyautogui.click()
    time.sleep(0.5)

# Iterate the loop to read the cell values to call function to transfer them
for row in range(26,27):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):

        if col[row].value is None:
            pass
        else:
            # prints clan members name
            if "X:" not in col[row].value:
                print("\n")
                print("**********")
                print(col[row].value)
                clannie = col[row].value
                time.sleep(2)
                # clannie = "Xan"
                print("**********")
            else:  # prints targets name and coordinates
                print(col[row].value)
                coords = col[row].value.upper().split("X:")[1:2]
                x_coord = coords[0].split("Y:")[0]
                y_coord = coords[0].split("Y:")[1]
                print(x_coord)
                print(y_coord)
                time.sleep(1)
                transfer_coords(x_coord, y_coord, clannie)

            # if "@" in col[row].value:
            #     # print("~~~~~~~~~~")
            #     label = col[row].value.split("@")[1:2]
            #     send_label(label[0])
            #     print(label[0])
            #     # print("~~~~~~~~~~")
            # elif "X:" not in col[row].value:
            #     print("**********")
            #     print(col[row].value)
            #     print("**********")
            #     clannie = col[row].value
            # else:
            #     print(col[row].value)
            #     coords = col[row].value.upper().split("X:")[1:2]
            #     print("coords", coords)
            #     x_coord = coords[0].split("Y:")[0]
            #     y_coord = coords[0].split("Y:")[1]
            #     print("x_coord", x_coord)
            #     print("y_coord", y_coord)
            #     time.sleep(1)
            #     transfer_coords(x_coord, y_coord, clannie)
